import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook

# Common date formats to check against
DATE_FORMATS = [
    "%Y-%m-%d", "%m-%d-%Y", "%d/%m/%Y", "%b %d, %Y", "%B %d, %Y",
    "%d-%b-%Y", "%d %B %Y", "%Y-%b-%d", "%m/%d/%Y", "%d %b, %Y",
    "%d %B, %Y", "%d %b %Y", "%d-%m-%Y", "%d/%m/%Y"
]

# The target format we want all dates converted to
OUTPUT_FORMAT = "%Y年%m月%d日"

# Day names in Chinese (used optionally)
CHINESE_WEEKDAYS = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']

# Convert a raw date string into the standard format
def normalize_date(date_str):
    if not isinstance(date_str, str):
        return date_str

    # Replace Chinese characters representing digits with standard numerals
    chinese_digits = {
        '〇': '0', '零': '0', '一': '1', '二': '2', '三': '3', '四': '4',
        '五': '5', '六': '6', '七': '7', '八': '8', '九': '9',
        '壹': '1', '贰': '2', '叁': '3', '肆': '4', '伍': '5',
        '陆': '6', '柒': '7', '捌': '8', '玖': '9'
    }

    cleaned = date_str.strip()
    cleaned = re.sub(r"\s+", "", cleaned)

    for zh_char, digit in chinese_digits.items():
        cleaned = cleaned.replace(zh_char, digit)

    cleaned = re.sub(r"[,\.\(\)]", "", cleaned)

    # Convert Chinese date format (e.g., 2023年8月6日)
    def chinese_to_date(m):
        y, mth, d = m.groups()
        try:
            dt = datetime(int(y), int(mth), int(d))
            formatted = dt.strftime(OUTPUT_FORMAT)
            # Convert digits back to Chinese numerals
            return ''.join({
                '0': '〇','1': '一','2': '二','3': '三','4': '四',
                '5': '五','6': '六','7': '七','8': '八','9': '九'
            }.get(ch, ch) for ch in formatted)
        except:
            return m.group(0)

    cleaned = re.sub(r"(\d{4})年(\d{1,2})月(\d{1,2})日", chinese_to_date, cleaned)

    # If not Chinese-style, try known formats
    for fmt in DATE_FORMATS:
        try:
            parsed = datetime.strptime(cleaned, fmt)
            return parsed.strftime(OUTPUT_FORMAT)
        except:
            continue

    return cleaned  # If nothing works, return as-is

# Apply normalization to individual cells
def replace_dates_in_text(cell):
    if isinstance(cell, datetime):
        return cell.strftime(OUTPUT_FORMAT)
    if not isinstance(cell, str):
        return cell
    return normalize_date(cell)

# Process an Excel file, normalize all date entries, and save to CSV
def normalize_dates_in_file(input_file, output_file):
    encoding = 'utf-8-sig'
    try:
        wb = load_workbook(input_file, read_only=True, data_only=True)
        ws = wb.active
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        df = pd.DataFrame(rows[1:], columns=rows[0])

        for col in df.columns:
            df[col] = df[col].apply(replace_dates_in_text)

        df.to_csv(output_file, index=False, encoding=encoding)
        print(f"✔ File saved as: {output_file}")
    except Exception as e:
        print(f"✘ Something went wrong: {e}")

# File paths to read from and write to
input_file = "/content/Normalized_Chinese_Long.xlsx"
output_file = "Chinese_Lunar.csv"
normalize_dates_in_file(input_file, output_file)
